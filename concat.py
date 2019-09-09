# -*- coding utf-8 -*- #

from numpy import *
import pickle
import xlrd
import os
from openpyxl import Workbook
from utils import mysql_util

# 下面这些变量需要您根据自己的具体情况选择
biaotou = ['Period', 'VehicleKey', 'MakeCode', 'Manufacturer', 'ManufacturerChinese', 'FamilyCode',
           'FamilyDescriptionChinese', 'VehicleTypeCode', 'YearGroup', 'MonthGroup', 'SequenceNum', 'Description',
           'DescriptionChinese', 'CurrentRelease', 'ImportStatus', 'LimitedEdition', 'Series', 'SeriesChinese',
           'BadgeDescription', 'BadgeDescriptionChinese', 'BadgeSecondaryDescription',
           'BadgeSecondaryDescriptionChinese', 'BodyStyleDescription', 'BodyStyleDescriptionChinese',
           'DriveDescription', 'DriveDescriptionChinese', 'GearTypeDescription', 'GearTypeDescriptionChinese',
           'EngineSize', 'EngineDescription', 'FuelTypeDescription', 'GrossVehicleMass', 'WheelBase', 'Height',
           'Length', 'Width', 'KerbWeight', 'Power', 'BuildCountryOriginDescription', 'AverageKM', 'GoodKM',
           'AvgWholesale', 'AvgRetail', 'GoodWholesale', 'GoodRetail', 'NewPrice', 'GoodWhsRV', 'GoodRtlRV', 'Y&M',
           'Y&M', 'Rolling系数', 'Rolling Rtl', 'Rolling系数1价格', 'Rolling系数2车体', 'Rolling Whs']

make_code = ['AUDI', 'MERC', 'BMW', 'PORS', 'JAGU', 'TESL']
# 在哪里搜索多个表格
filelocation = "./data1"
# 当前文件夹下搜索的文件名后缀
fileform = "xlsx"
# 将合并后的表格存放到的位置
filedestination = "./data1"
# 合并后的表格命名为file
file = "result1"
# 数据保存pkl的位置
pkl_file_name = 'matrix_1.pkl'

# 首先查找默认文件夹下有多少文档需要整合
file_array = os.listdir('./data1')
# 以上是从pythonscripts文件夹下读取所有excel表格，file_array
print("在默认文件夹下有%d个文档哦" % len(file_array))
ge = len(file_array)
matrix = [None] * ge

db = mysql_util.get_db(host='localhost', user='root', password='root1234', port=3306, db='redbook')
sql_year_mouth_group = 'select ratio from year_mouth_group_ratio where year_mouth_group = %s and mouth = %s'


# 实现读写数据
# 下面是将所有文件读数据到三维列表cell[][][]中（不包含表头）
sum = 0
load_file = False
if load_file:
    # 加载保存好的数据
    with open(pkl_file_name, 'rb') as pkl_file:
        matrix = pickle.load(pkl_file)
else:
    for i in range(ge):
        file_name = filelocation + '/' + file_array[i]
        wb = xlrd.open_workbook(file_name)
        sheet = wb.sheet_by_index(0)
        nrows = sheet.nrows

        # 计算出需要挑选的行数
        now_tows = 0
        for j in range(1, nrows):
            mackCode = sheet.cell(j, 1).value
            if mackCode in make_code:
                now_tows += 1
        matrix[i] = [0] * (now_tows)

        # 计算出需要挑选的列数
        ncols = sheet.ncols
        for m in range(now_tows):
            matrix[i][m] = [""] * len(biaotou)

        # 将当前文件中需要挑选的列小标保存起来
        title_index = {}
        for k in range(0, ncols):
            title = sheet.cell(0, k).value
            if title in biaotou:
                title_index[k] = title

        # 将数据保存到matrix 数组中
        indexs = title_index.keys()
        num = 0
        year_mouth = file_array[i][2:-5]
        for j in range(1, nrows):
            mackCode = sheet.cell(j, 1).value
            if mackCode in make_code:
                matrix[i][num][0] = year_mouth
                for k in range(0, ncols):
                    if k in indexs:
                        index = biaotou.index(title_index[k])
                        matrix[i][num][index] = sheet.cell(j, k).value
                num += 1
        print(num, i)
        sum += num
        # break
    print(sum)

    # 防止保存的代码有问题，将抽取出来的数据保存起来
    with open(pkl_file_name, 'wb') as pkl_file:
        pickle.dump(matrix, pkl_file)

# 文件合并
workbook = Workbook()
booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet

# 下面是把表头写上
for i in range(0, len(biaotou)):
    booksheet.cell(1, i + 1).value = biaotou[i]

zh = 2
for i in range(ge):

    # if i > 0 and i % 6 == 0:
    #     print("我已经将%d个文件合并成1个文件，并命名为%s.xls.快打开看看正确不？" % (ge, file + str(i)))
    #     workbook.save(filedestination + '/' + file + str(i) + ".xlsx")
    #
    #     workbook = Workbook()
    #     booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet
    #
    #     # 下面是把表头写上
    #     for w in range(0, len(biaotou)):
    #         booksheet.cell(1, w + 1).value = biaotou[w]

    for j in range(len(matrix[i])):
        for k in range(len(matrix[i][j])):
            booksheet.cell(zh, k + 1).value = matrix[i][j][k]
        zh = zh + 1
print("我已经将%d个文件合并成1个文件，并命名为%s.xlsx.快打开看看正确不？" % (ge, file))
workbook.save(filedestination + "/" + file + ".xlsx")
